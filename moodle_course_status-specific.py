#!/usr/bin/env python3
import argparse
import time
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import re
import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ----------------- Settings -----------------
DEFAULT_MAX_THREADS = 8
REQUEST_TIMEOUT = 60
RETRY_COUNT = 3
RETRY_SLEEP_SECONDS = 1.5
# --------------------------------------------

import re

def sanitize_sheet_name(name):
    """
    Sanitize Excel sheet names by:
    - Removing spaces
    - Removing underscores, hyphens, brackets, braces
    - Removing all special characters (keep only A–Z, a–z, 0–9)
    - Limiting to 31 characters (Excel rule)
    """
    # Keep only letters and digits
    cleaned = re.sub(r'[^A-Za-z0-9]', '', name)

    # Excel sheet names max length = 31
    return cleaned[:31]

def call_moodle(url, token, function, params=None):
    """Generic Moodle REST call with retries."""
    if params is None:
        params = {}
    full = f"{url}/webservice/rest/server.php"
    base = {
        "wstoken": token,
        "wsfunction": function,
        "moodlewsrestformat": "json"
    }
    for attempt in range(1, RETRY_COUNT + 1):
        try:
            r = requests.post(full, params=base, data=params, timeout=REQUEST_TIMEOUT)
            r.raise_for_status()
            data = r.json()
            if isinstance(data, dict) and data.get("exception"):
                raise RuntimeError(f"Moodle exception: {data.get('message')} | {data.get('errorcode')}")
            return data
        except Exception as e:
            if attempt == RETRY_COUNT:
                raise
            time.sleep(RETRY_SLEEP_SECONDS * attempt)
    return None

def get_all_courses(url, token, specific_courses=None):
    data = call_moodle(url, token, "core_course_get_courses_by_field")
    courses = data.get("courses", [])
    if specific_courses:
        specific_lower = [c.lower() for c in specific_courses]
        filtered = []
        for c in courses:
            if str(c["id"]) in specific_lower or c["fullname"].lower() in specific_lower:
                filtered.append(c)
        return filtered
    return courses

def get_course_users(url, token, course_id):
    return call_moodle(url, token, "core_enrol_get_enrolled_users", {"courseid": course_id})

def get_course_activities(url, token, course_id):
    sections = call_moodle(url, token, "core_course_get_contents", {"courseid": course_id})
    activities = {}
    for sec in sections:
        for mod in sec.get("modules", []):
            if mod.get("uservisible", True) and not mod.get("deletioninprogress", False):
                cmid = mod.get("id")
                activities[cmid] = {"name": mod.get("name"), "modname": mod.get("modname")}
    return activities

def get_course_completion(url, token, course_id, user_id):
    try:
        resp = call_moodle(url, token, "core_completion_get_course_completion_status",
                           {"courseid": course_id, "userid": user_id})
        if isinstance(resp, dict) and "completionstatus" in resp:
            completed = resp["completionstatus"].get("completed", None)
            return 100 if completed else 0, "Completed" if completed else "Incomplete"
    except Exception:
        pass
    return 0, "N/A"

def get_activity_completion(url, token, course_id, user_id):
    out = {}
    try:
        data = call_moodle(url, token, "core_completion_get_activities_completion_status",
                           {"courseid": course_id, "userid": user_id})
        for s in data.get("statuses", []):
            cmid = s.get("cmid")
            state = s.get("state", 0)
            if state in (1, 2):
                status = "Completed"
            elif state == 3:
                status = "Failed"
            else:
                status = "Incomplete"
            out[cmid] = status
    except Exception:
        pass
    return out

def role_names(user):
    try:
        return ", ".join([r.get("name", str(r.get("roleid", ""))) for r in user.get("roles", [])])
    except Exception:
        return ""

def last_access_str(user):
    ts = user.get("lastaccess")
    return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S") if ts else ""

def flatten_custom_fields(user):
    d = {}
    for f in user.get("customfields", []) or []:
        key = f.get("shortname") or f.get("name")
        d[key] = f.get("value", "")
    return d

def process_one_course(course, url, token):
    cid = course.get("id")
    cname = course.get("fullname", "")
    cshort = course.get("shortname", "")
    ccat = course.get("categoryid", "")

    print(f"Processing course: {cname} (ID={cid})")

    activities = get_course_activities(url, token, cid)
    activity_columns = []
    for cmid, meta in activities.items():
        col = f"{meta['modname'].capitalize()}: {meta['name']}"
        activity_columns.append((cmid, col))

    try:
        users = get_course_users(url, token, cid)
    except Exception as e:
        print(f"  ❌ Failed to get users for course {cname}: {e}")
        return [], [], [], [col for _, col in activity_columns]

    per_course_rows = []
    consolidated_rows = []
    enrollment_rows = []

    for u in users:
        uid = u.get("id")
        base_user = {
            "User ID": uid,
            "Full Name": u.get("fullname", ""),
            "Username": u.get("username", ""),
            "Email": u.get("email", ""),
            "Department": u.get("department", ""),
            "Institution": u.get("institution", ""),
            "City": u.get("city", ""),
            "Country": u.get("country", ""),
            "Last Access": last_access_str(u),
            "Role(s)": role_names(u),
        }
        base_user.update(flatten_custom_fields(u))

        c_pct, c_status = get_course_completion(url, token, cid, uid)
        a_map = get_activity_completion(url, token, cid, uid)

        # Per-course row (keep all activity columns)
        course_row = {
            "Course ID": cid,
            "Course Name": cname,
            "Course Shortname": cshort,
            **base_user,
            "Course Completion Status": c_status,
            "Completion %": c_pct
        }
        for cmid, colname in activity_columns:
            course_row[colname] = a_map.get(cmid, "Incomplete")
        per_course_rows.append(course_row)

        # Consolidated row (only selected columns)
        consolidated_row = [
            uid,
            u.get("fullname", ""),
            "",  # Manager placeholder
            u.get("email", ""),
            cid,
            cname,
            last_access_str(u),
            role_names(u),
            c_pct,
            c_status
        ]
        consolidated_rows.append(consolidated_row)

        enrollment_rows.append({
            "User ID": uid,
            "Full Name": u.get("fullname", ""),
            "Username": u.get("username", ""),
            "Email": u.get("email", ""),
            "Course ID": cid,
            "Course Name": cname,
            "Course Shortname": cshort,
            "Role(s)": role_names(u),
            "Last Access": last_access_str(u)
        })

    return per_course_rows, consolidated_rows, enrollment_rows, [col for _, col in activity_columns]

def autosize_columns(ws):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0]
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)

def main():
    ap = argparse.ArgumentParser(description="Moodle: per-activity completion report (single or all courses)")
    ap.add_argument("--url", required=True, help="Moodle base URL")
    ap.add_argument("--token", required=True, help="Moodle web service token")
    ap.add_argument("--threads", type=int, default=DEFAULT_MAX_THREADS, help="Max concurrent courses")
    ap.add_argument("--courseid", type=int, help="If provided, only this course ID will be processed")
    ap.add_argument("--courses_file", help="File with list of course IDs or names, one per line")
    args = ap.parse_args()

    url = args.url.rstrip("/")
    token = args.token
    max_threads = max(1, args.threads)

    specific_courses = None
    if args.courses_file:
        with open(args.courses_file) as f:
            specific_courses = [line.strip() for line in f if line.strip()]

    if args.courseid:
        courses = get_all_courses(url, token, [str(args.courseid)])
        print(f"Processing only course ID={args.courseid} ({len(courses)} found)")
    else:
        courses = get_all_courses(url, token, specific_courses)
        print(f"Found {len(courses)} courses.")

    consolidated_all, enrollment_all, per_course_results = [], [], []

    with ThreadPoolExecutor(max_workers=max_threads) as ex:
        fut_map = {ex.submit(process_one_course, c, url, token): c for c in courses}
        for fut in as_completed(fut_map):
            course = fut_map[fut]
            cname = course.get("fullname", "")
            cshort = course.get("shortname", f"course_{course.get('id')}")
            try:
                per_course_rows, consolidated_rows, enrollment_rows, act_cols = fut.result()
                per_course_results.append((cshort, per_course_rows, act_cols))
                consolidated_all.extend(consolidated_rows)
                enrollment_all.extend(enrollment_rows)
                print(f"  ✅ Done: {cname} ({len(per_course_rows)} user rows)")
            except Exception as e:
                print(f"  ❌ Error processing {cname}: {e}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Moodle_Completion_Report_{ts}.xlsx"
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "All Courses – Consolidated"

    # Only requested columns in consolidated sheet
    main_headers = ["User ID","Full Name","Manager","Email","Course ID","Course Name",
                    "Last Access","Role(s)","Completion %","Course Completion Status"]
    ws_main.append(main_headers)
    for row in consolidated_all:
        ws_main.append(row)
    autosize_columns(ws_main)

    # Per-course sheets with all activities as before
    for cshort, rows, act_cols in sorted(per_course_results, key=lambda x: x[0].lower()):
        ws = wb.create_sheet(title=sanitize_sheet_name(cshort) or "Course")
        base = [
            "Course ID","Course Name","Course Shortname",
            "User ID","Full Name","Username","Email",
            "Department","Institution","City","Country","Last Access","Role(s)"
        ]
        trailing = ["Completion %","Course Completion Status"]
        headers = base + list(act_cols) + trailing
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        autosize_columns(ws)

    # Enrollment sheet
    ws_en = wb.create_sheet(title="Enrollments")
    en_headers = ["User ID","Full Name","Username","Email","Course ID","Course Name","Course Shortname","Role(s)","Last Access"]
    ws_en.append(en_headers)
    for r in enrollment_all:
        ws_en.append([r.get(h, "") for h in en_headers])
    autosize_columns(ws_en)

    wb.save(filename)
    print(f"\n✅ Report saved: {filename}")

if __name__ == "__main__":
    main()

